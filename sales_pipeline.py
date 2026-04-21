"""
Sales Analytics & Incentive Optimization Pipeline
===================================================
Transforms raw retail transaction data into:
  - Agent-level KPI analysis
  - Monthly performance tracking
  - Incentive calculation (2 schemes: Old vs New)
  - Optimization comparison
Output: sales_analytics_report.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import random

random.seed(42)
np.random.seed(42)

# ─────────────────────────────────────────────
# 1. LOAD & CLEAN DATA
# ─────────────────────────────────────────────
print("📥 Loading data...")
df = pd.read_csv("retail_sales_dataset.csv")
df["Date"] = pd.to_datetime(df["Date"])
df = df[df["Date"].dt.year == 2023].copy()  # keep 12 months clean
df["Month"] = df["Date"].dt.to_period("M").astype(str)
df["Quarter"] = "Q" + df["Date"].dt.quarter.astype(str)

# ─────────────────────────────────────────────
# 2. SIMULATE AGENTS & REGIONS
# ─────────────────────────────────────────────
print("🧑‍💼 Assigning sales agents...")

AGENTS = [
    "An Nguyen", "Binh Tran", "Chi Le", "Duc Pham", "Em Hoang",
    "Fong Vu",   "Giang Do",  "Hoa Bui", "Iris Ngo", "Jack Dang"
]
REGIONS = {
    "An Nguyen": "North",  "Binh Tran": "North",
    "Chi Le":    "Central","Duc Pham":  "Central",
    "Em Hoang":  "South",  "Fong Vu":   "South",
    "Giang Do":  "South",  "Hoa Bui":   "East",
    "Iris Ngo":  "East",   "Jack Dang": "East"
}
CATEGORIES = {
    "An Nguyen": "Electronics", "Binh Tran": "Clothing",
    "Chi Le":    "Beauty",      "Duc Pham":  "Electronics",
    "Em Hoang":  "Clothing",    "Fong Vu":   "Beauty",
    "Giang Do":  "Electronics", "Hoa Bui":   "Clothing",
    "Iris Ngo":  "Beauty",      "Jack Dang": "Electronics"
}

# Assign agent — weighted so each has ~100 transactions
agent_pool = AGENTS * (len(df) // len(AGENTS) + 1)
df["Agent"] = random.choices(AGENTS, k=len(df))
df["Region"] = df["Agent"].map(REGIONS)
df["Agent_Category"] = df["Agent"].map(CATEGORIES)

# ─────────────────────────────────────────────
# 3. GENERATE TARGETS (monthly per agent)
# ─────────────────────────────────────────────
print("🎯 Generating monthly targets...")

months = sorted(df["Month"].unique())
agent_month_sales = (
    df.groupby(["Agent", "Month"])["Total Amount"]
    .sum().reset_index()
    .rename(columns={"Total Amount": "Actual_Sales"})
)

# Target = actual * random factor (so agents vary around 80–120%)
agent_month_sales["Target"] = (
    agent_month_sales["Actual_Sales"]
    / np.random.uniform(0.75, 1.30, size=len(agent_month_sales))
).round(-2)  # round to nearest 100

agent_month_sales["Pct_Target"] = (
    agent_month_sales["Actual_Sales"] / agent_month_sales["Target"]
).round(4)

# ─────────────────────────────────────────────
# 4. KPI CALCULATIONS
# ─────────────────────────────────────────────
print("📊 Computing KPIs...")

# -- Monthly KPI per agent
monthly_kpi = agent_month_sales.copy()
monthly_kpi["Region"] = monthly_kpi["Agent"].map(REGIONS)
monthly_kpi["Quarter"] = pd.PeriodIndex(monthly_kpi["Month"], freq="M").to_timestamp().to_series().reset_index(drop=True)
monthly_kpi["Quarter"] = pd.to_datetime(monthly_kpi["Month"]).dt.quarter.apply(lambda q: f"Q{q}")

# Rank within month
monthly_kpi["Monthly_Rank"] = (
    monthly_kpi.groupby("Month")["Actual_Sales"]
    .rank(ascending=False, method="min")
    .astype(int)
)

# ─────────────────────────────────────────────
# 5. INCENTIVE ENGINE
# ─────────────────────────────────────────────
print("💰 Calculating incentives...")

def calc_incentive_old(row):
    """
    OLD SCHEME:
      < 80% target  → 0%
      80–100%       → 5%
      > 100%        → 8%
    """
    pct = row["Pct_Target"]
    sales = row["Actual_Sales"]
    if pct < 0.80:
        rate = 0.00
    elif pct < 1.00:
        rate = 0.05
    else:
        rate = 0.08
    return round(sales * rate, 2)


def calc_incentive_new(row, top_agents):
    """
    NEW SCHEME:
      < 80% target  → 0%
      80–100%       → 5%
      > 100%        → 10%  (+2% vs old)
      Top 10% monthly ranking → +3% bonus on top
    """
    pct = row["Pct_Target"]
    sales = row["Actual_Sales"]
    if pct < 0.80:
        rate = 0.00
    elif pct < 1.00:
        rate = 0.05
    else:
        rate = 0.10
    # Top 10% bonus
    if row["Agent"] in top_agents.get(row["Month"], set()):
        rate += 0.03
    return round(sales * rate, 2)


# Identify top 10% agents per month
top_10pct = {}
n_top = max(1, round(len(AGENTS) * 0.10))  # 1 agent
for month, grp in monthly_kpi.groupby("Month"):
    top_10pct[month] = set(
        grp.nlargest(n_top, "Actual_Sales")["Agent"].values
    )

monthly_kpi["Incentive_Old"] = monthly_kpi.apply(calc_incentive_old, axis=1)
monthly_kpi["Incentive_New"] = monthly_kpi.apply(
    lambda r: calc_incentive_new(r, top_10pct), axis=1
)
monthly_kpi["Incentive_Diff"] = monthly_kpi["Incentive_New"] - monthly_kpi["Incentive_Old"]

# Performance tier label
def perf_tier(pct):
    if pct >= 1.20: return "Star"
    if pct >= 1.00: return "On Target"
    if pct >= 0.80: return "Near Miss"
    return "Below Target"

monthly_kpi["Performance_Tier"] = monthly_kpi["Pct_Target"].apply(perf_tier)

# ─────────────────────────────────────────────
# 6. AGENT SUMMARY (annual)
# ─────────────────────────────────────────────
print("📋 Building agent summary...")

agent_summary = monthly_kpi.groupby("Agent").agg(
    Region=("Region", "first"),
    Total_Sales=("Actual_Sales", "sum"),
    Total_Target=("Target", "sum"),
    Avg_Pct_Target=("Pct_Target", "mean"),
    Months_Above_Target=("Pct_Target", lambda x: (x >= 1.0).sum()),
    Months_Below_80=("Pct_Target", lambda x: (x < 0.8).sum()),
    Total_Incentive_Old=("Incentive_Old", "sum"),
    Total_Incentive_New=("Incentive_New", "sum"),
).reset_index()

agent_summary["Annual_Pct_Target"] = (
    agent_summary["Total_Sales"] / agent_summary["Total_Target"]
).round(4)

agent_summary["ROI_Old"] = (
    agent_summary["Total_Sales"] / agent_summary["Total_Incentive_Old"].replace(0, np.nan)
).round(2)
agent_summary["ROI_New"] = (
    agent_summary["Total_Sales"] / agent_summary["Total_Incentive_New"].replace(0, np.nan)
).round(2)

agent_summary["Incentive_Change"] = (
    agent_summary["Total_Incentive_New"] - agent_summary["Total_Incentive_Old"]
).round(2)

agent_summary["Incentive_Change_Pct"] = (
    agent_summary["Incentive_Change"] / agent_summary["Total_Incentive_Old"].replace(0, np.nan)
).round(4)

# Rank agents
agent_summary["Annual_Rank"] = (
    agent_summary["Total_Sales"].rank(ascending=False, method="min").astype(int)
)

# Overpaid / Underpaid flag
median_roi = agent_summary["ROI_New"].median()
agent_summary["Pay_Status"] = agent_summary.apply(
    lambda r: "Overpaid" if (r["ROI_New"] < median_roi * 0.85 and r["Annual_Pct_Target"] < 1.0)
    else ("Underpaid" if (r["ROI_New"] > median_roi * 1.15 and r["Annual_Pct_Target"] >= 1.0)
          else "Fair"),
    axis=1
)

# ─────────────────────────────────────────────
# 7. SCHEME COMPARISON SUMMARY
# ─────────────────────────────────────────────
print("🔄 Building scheme comparison...")

scheme_comparison = pd.DataFrame({
    "Metric": [
        "Total Incentive Cost",
        "Total Sales",
        "Overall ROI (Sales / Cost)",
        "Avg % Target Achieved",
        "Agents Above 100% Target",
        "Agents Below 80% Target",
    ],
    "Old Scheme": [
        monthly_kpi["Incentive_Old"].sum(),
        monthly_kpi["Actual_Sales"].sum(),
        round(monthly_kpi["Actual_Sales"].sum() / monthly_kpi["Incentive_Old"].sum(), 2),
        round(monthly_kpi["Pct_Target"].mean(), 4),
        (monthly_kpi["Pct_Target"] >= 1.0).sum(),
        (monthly_kpi["Pct_Target"] < 0.8).sum(),
    ],
    "New Scheme": [
        monthly_kpi["Incentive_New"].sum(),
        monthly_kpi["Actual_Sales"].sum(),
        round(monthly_kpi["Actual_Sales"].sum() / monthly_kpi["Incentive_New"].sum(), 2),
        round(monthly_kpi["Pct_Target"].mean(), 4),
        (monthly_kpi["Pct_Target"] >= 1.0).sum(),
        (monthly_kpi["Pct_Target"] < 0.8).sum(),
    ]
})

# ─────────────────────────────────────────────
# 8. REGION SUMMARY
# ─────────────────────────────────────────────
region_summary = monthly_kpi.groupby("Region").agg(
    Total_Sales=("Actual_Sales", "sum"),
    Total_Target=("Target", "sum"),
    Total_Incentive_Old=("Incentive_Old", "sum"),
    Total_Incentive_New=("Incentive_New", "sum"),
    Num_Agents=("Agent", "nunique"),
).reset_index()
region_summary["Pct_Target"] = (region_summary["Total_Sales"] / region_summary["Total_Target"]).round(4)
region_summary["ROI_New"] = (region_summary["Total_Sales"] / region_summary["Total_Incentive_New"]).round(2)

# ─────────────────────────────────────────────
# 9. TRANSACTION DATA ENRICHED
# ─────────────────────────────────────────────
transactions = df.merge(
    monthly_kpi[["Agent", "Month", "Pct_Target", "Performance_Tier",
                 "Incentive_Old", "Incentive_New", "Monthly_Rank"]],
    on=["Agent", "Month"],
    how="left"
)

# ─────────────────────────────────────────────
# 10. EXPORT TO EXCEL
# ─────────────────────────────────────────────
print("📁 Writing Excel file...")

OUTPUT = "sales_analytics_report.xlsx"

# Color palette
C_HEADER_BG   = "1F3864"   # deep navy
C_HEADER_FG   = "FFFFFF"
C_SUBHDR_BG   = "2E75B6"   # blue
C_SUBHDR_FG   = "FFFFFF"
C_ALT_ROW     = "EBF3FB"   # light blue
C_WHITE       = "FFFFFF"
C_GREEN_BG    = "E2EFDA"
C_RED_BG      = "FCE4D6"
C_YELLOW_BG   = "FFF2CC"
C_STAR        = "FFD700"

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(size=11, bold=True, color=C_HEADER_FG):
    return Font(name="Arial", size=size, bold=bold, color=color)

def cell_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)

def hdr_fill(color=C_HEADER_BG):
    return PatternFill("solid", start_color=color)

def row_fill(color):
    return PatternFill("solid", start_color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center")

def pct_fmt(val):
    return f"{val:.1%}"

def num_fmt(val):
    return f"{val:,.0f}"

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ── Helper: style a header row ──────────────────
def style_header_row(ws, row_num, col_start, col_end, bg=C_HEADER_BG):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = hdr_font(color=C_HEADER_FG)
        cell.fill = hdr_fill(bg)
        cell.alignment = center()
        cell.border = border

def style_data_row(ws, row_num, col_start, col_end, alt=False):
    bg = C_ALT_ROW if alt else C_WHITE
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = cell_font()
        cell.fill = row_fill(bg)
        cell.alignment = left()
        cell.border = border

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

# ══════════════════════════════════════════════
# SHEET 1: Agent Summary
# ══════════════════════════════════════════════
ws1 = wb.create_sheet("Agent Summary")
ws1.freeze_panes = "A3"

title_cell = ws1.cell(1, 1, "🏆  AGENT ANNUAL PERFORMANCE SUMMARY")
title_cell.font = Font(name="Arial", size=14, bold=True, color=C_HEADER_BG)
title_cell.alignment = left()
ws1.merge_cells("A1:P1")

cols = [
    "Rank", "Agent", "Region",
    "Total Sales", "Total Target", "% Target",
    "Months ≥100%", "Months <80%",
    "Incentive (Old)", "Incentive (New)", "Incentive Δ", "Incentive Δ%",
    "ROI (Old)", "ROI (New)",
    "Performance Tier", "Pay Status"
]
for c, col in enumerate(cols, 1):
    cell = ws1.cell(2, c, col)
    cell.font = hdr_font(size=10)
    cell.fill = hdr_fill(C_HEADER_BG)
    cell.alignment = center()
    cell.border = border

agent_summary_sorted = agent_summary.sort_values("Annual_Rank")
for r, row in enumerate(agent_summary_sorted.itertuples(), 3):
    alt = (r % 2 == 0)
    style_data_row(ws1, r, 1, len(cols), alt)

    ws1.cell(r, 1, row.Annual_Rank)
    ws1.cell(r, 2, row.Agent)
    ws1.cell(r, 3, row.Region)
    ws1.cell(r, 4, row.Total_Sales).number_format = "#,##0"
    ws1.cell(r, 5, row.Total_Target).number_format = "#,##0"
    ws1.cell(r, 6, row.Annual_Pct_Target).number_format = "0.0%"
    ws1.cell(r, 7, row.Months_Above_Target)
    ws1.cell(r, 8, row.Months_Below_80)
    ws1.cell(r, 9, row.Total_Incentive_Old).number_format = "#,##0"
    ws1.cell(r, 10, row.Total_Incentive_New).number_format = "#,##0"
    ws1.cell(r, 11, row.Incentive_Change).number_format = "#,##0"
    ws1.cell(r, 12, row.Incentive_Change_Pct).number_format = "0.0%"
    ws1.cell(r, 13, row.ROI_Old).number_format = "0.00"
    ws1.cell(r, 14, row.ROI_New).number_format = "0.00"
    ws1.cell(r, 15, row.Annual_Pct_Target)
    ws1.cell(r, 16, row.Pay_Status)

    # Conditional color: % Target
    pct_cell = ws1.cell(r, 6)
    if row.Annual_Pct_Target >= 1.0:
        pct_cell.fill = row_fill(C_GREEN_BG)
    elif row.Annual_Pct_Target < 0.80:
        pct_cell.fill = row_fill(C_RED_BG)

    # Conditional color: Pay Status
    pay_cell = ws1.cell(r, 16)
    if row.Pay_Status == "Overpaid":
        pay_cell.fill = row_fill(C_RED_BG)
        pay_cell.font = Font(name="Arial", size=10, bold=True, color="C00000")
    elif row.Pay_Status == "Underpaid":
        pay_cell.fill = row_fill(C_YELLOW_BG)
        pay_cell.font = Font(name="Arial", size=10, bold=True, color="7F6000")

    # Performance tier col (remove formula, just color)
    tier_cell = ws1.cell(r, 15)
    tier_val = row.Annual_Pct_Target
    if tier_val >= 1.20:
        tier_cell.value = "⭐ Star"
        tier_cell.fill = row_fill(C_STAR)
    elif tier_val >= 1.00:
        tier_cell.value = "✅ On Target"
        tier_cell.fill = row_fill(C_GREEN_BG)
    elif tier_val >= 0.80:
        tier_cell.value = "⚠️ Near Miss"
        tier_cell.fill = row_fill(C_YELLOW_BG)
    else:
        tier_cell.value = "❌ Below Target"
        tier_cell.fill = row_fill(C_RED_BG)

set_col_widths(ws1, [6, 16, 9, 13, 13, 10, 12, 12, 15, 15, 13, 13, 10, 10, 15, 13])

# ══════════════════════════════════════════════
# SHEET 2: Monthly KPI
# ══════════════════════════════════════════════
ws2 = wb.create_sheet("Monthly KPI")
ws2.freeze_panes = "A3"

title2 = ws2.cell(1, 1, "📅  MONTHLY KPI — ALL AGENTS")
title2.font = Font(name="Arial", size=14, bold=True, color=C_HEADER_BG)
ws2.merge_cells("A1:N1")

cols2 = [
    "Month", "Quarter", "Agent", "Region",
    "Actual Sales", "Target", "% Target",
    "Monthly Rank", "Performance Tier",
    "Incentive (Old)", "Incentive (New)", "Incentive Δ",
    "Overpaid / Underpaid", "Notes"
]
for c, col in enumerate(cols2, 1):
    cell = ws2.cell(2, c, col)
    cell.font = hdr_font(size=10)
    cell.fill = hdr_fill(C_HEADER_BG)
    cell.alignment = center()
    cell.border = border

monthly_sorted = monthly_kpi.sort_values(["Month", "Monthly_Rank"])

for r, row in enumerate(monthly_sorted.itertuples(), 3):
    alt = (r % 2 == 0)
    style_data_row(ws2, r, 1, len(cols2), alt)

    ws2.cell(r, 1, row.Month)
    ws2.cell(r, 2, row.Quarter)
    ws2.cell(r, 3, row.Agent)
    ws2.cell(r, 4, row.Region)
    ws2.cell(r, 5, row.Actual_Sales).number_format = "#,##0"
    ws2.cell(r, 6, row.Target).number_format = "#,##0"
    pct_c = ws2.cell(r, 7, row.Pct_Target)
    pct_c.number_format = "0.0%"
    ws2.cell(r, 8, row.Monthly_Rank)
    ws2.cell(r, 9, row.Performance_Tier)
    ws2.cell(r, 10, row.Incentive_Old).number_format = "#,##0"
    ws2.cell(r, 11, row.Incentive_New).number_format = "#,##0"
    diff_c = ws2.cell(r, 12, row.Incentive_Diff)
    diff_c.number_format = "#,##0"
    ws2.cell(r, 13, "")  # placeholder

    # Color % Target
    if row.Pct_Target >= 1.0:
        pct_c.fill = row_fill(C_GREEN_BG)
    elif row.Pct_Target < 0.80:
        pct_c.fill = row_fill(C_RED_BG)

    # Color diff
    if row.Incentive_Diff > 0:
        diff_c.fill = row_fill(C_GREEN_BG)

set_col_widths(ws2, [9, 8, 16, 9, 13, 13, 10, 12, 14, 15, 15, 12, 16, 20])

# ══════════════════════════════════════════════
# SHEET 3: Scheme Comparison
# ══════════════════════════════════════════════
ws3 = wb.create_sheet("Scheme Comparison")

title3 = ws3.cell(1, 1, "🔄  INCENTIVE SCHEME COMPARISON — OLD vs NEW")
title3.font = Font(name="Arial", size=14, bold=True, color=C_HEADER_BG)
ws3.merge_cells("A1:E1")

# Header
for c, h in enumerate(["Metric", "Old Scheme", "New Scheme", "Δ Change", "Verdict"], 1):
    cell = ws3.cell(2, c, h)
    cell.font = hdr_font(size=11)
    cell.fill = hdr_fill(C_SUBHDR_BG)
    cell.alignment = center()
    cell.border = border

for r, row in enumerate(scheme_comparison.itertuples(), 3):
    alt = (r % 2 == 0)
    bg = C_ALT_ROW if alt else C_WHITE

    for c in range(1, 6):
        ws3.cell(r, c).fill = row_fill(bg)
        ws3.cell(r, c).font = cell_font()
        ws3.cell(r, c).alignment = left()
        ws3.cell(r, c).border = border

    ws3.cell(r, 1, row.Metric)
    old_val = row._2
    new_val = row._3

    if r == 3:   # Incentive cost
        ws3.cell(r, 2, old_val).number_format = "#,##0"
        ws3.cell(r, 3, new_val).number_format = "#,##0"
        delta = new_val - old_val
        ws3.cell(r, 4, delta).number_format = "+#,##0;-#,##0"
        ws3.cell(r, 5, "New costs more" if delta > 0 else "New saves money")
    elif r == 4:  # Total Sales
        ws3.cell(r, 2, old_val).number_format = "#,##0"
        ws3.cell(r, 3, new_val).number_format = "#,##0"
        ws3.cell(r, 4, "–")
        ws3.cell(r, 5, "Same base data")
    elif r == 5:  # ROI
        ws3.cell(r, 2, old_val).number_format = "0.00"
        ws3.cell(r, 3, new_val).number_format = "0.00"
        delta = new_val - old_val
        ws3.cell(r, 4, delta).number_format = "+0.00;-0.00"
        ws3.cell(r, 5, "✅ New scheme better ROI" if new_val > old_val else "⚠️ Old scheme better ROI")
    elif r == 6:  # Avg % target
        ws3.cell(r, 2, old_val).number_format = "0.0%"
        ws3.cell(r, 3, new_val).number_format = "0.0%"
        ws3.cell(r, 4, "–")
        ws3.cell(r, 5, "–")
    else:
        ws3.cell(r, 2, old_val)
        ws3.cell(r, 3, new_val)
        ws3.cell(r, 4, new_val - old_val)
        ws3.cell(r, 5, "–")

# Incentive rules legend
ws3.cell(10, 1, "INCENTIVE RULES REFERENCE").font = hdr_font(12, color=C_HEADER_BG)
ws3.cell(10, 1).fill = row_fill("DCE6F1")
ws3.merge_cells("A10:E10")

rules = [
    ["Tier", "% Target Achieved", "Old Scheme Rate", "New Scheme Rate", "Notes"],
    ["Below Threshold", "< 80%",   "0%",  "0%",   "No incentive"],
    ["Standard",        "80–99%",  "5%",  "5%",   "Same both schemes"],
    ["Above Target",    "≥ 100%",  "8%",  "10%",  "+2% uplift in New"],
    ["Top 10% Bonus",   "Any tier","–",   "+3%",  "Applied on top of base rate"],
]
for rr, rule_row in enumerate(rules, 11):
    for cc, val in enumerate(rule_row, 1):
        c = ws3.cell(rr, cc, val)
        c.border = border
        if rr == 11:
            c.font = hdr_font(10, color=C_HEADER_FG)
            c.fill = hdr_fill(C_SUBHDR_BG)
            c.alignment = center()
        else:
            c.font = cell_font()
            c.fill = row_fill(C_ALT_ROW if rr % 2 == 0 else C_WHITE)
            c.alignment = left()

set_col_widths(ws3, [28, 22, 16, 16, 30])

# ══════════════════════════════════════════════
# SHEET 4: Region Summary
# ══════════════════════════════════════════════
ws4 = wb.create_sheet("Region Summary")

title4 = ws4.cell(1, 1, "🗺️  REGIONAL PERFORMANCE SUMMARY")
title4.font = Font(name="Arial", size=14, bold=True, color=C_HEADER_BG)
ws4.merge_cells("A1:H1")

cols4 = ["Region", "Agents", "Total Sales", "Total Target", "% Target",
         "Incentive (Old)", "Incentive (New)", "ROI (New)"]
for c, h in enumerate(cols4, 1):
    cell = ws4.cell(2, c, h)
    cell.font = hdr_font(10)
    cell.fill = hdr_fill(C_HEADER_BG)
    cell.alignment = center()
    cell.border = border

for r, row in enumerate(region_summary.sort_values("Total_Sales", ascending=False).itertuples(), 3):
    alt = (r % 2 == 0)
    style_data_row(ws4, r, 1, len(cols4), alt)
    ws4.cell(r, 1, row.Region)
    ws4.cell(r, 2, row.Num_Agents)
    ws4.cell(r, 3, row.Total_Sales).number_format = "#,##0"
    ws4.cell(r, 4, row.Total_Target).number_format = "#,##0"
    pct = ws4.cell(r, 5, row.Pct_Target)
    pct.number_format = "0.0%"
    ws4.cell(r, 6, row.Total_Incentive_Old).number_format = "#,##0"
    ws4.cell(r, 7, row.Total_Incentive_New).number_format = "#,##0"
    ws4.cell(r, 8, row.ROI_New).number_format = "0.00"
    if row.Pct_Target >= 1.0:
        pct.fill = row_fill(C_GREEN_BG)
    elif row.Pct_Target < 0.80:
        pct.fill = row_fill(C_RED_BG)

set_col_widths(ws4, [12, 8, 14, 14, 10, 16, 16, 11])

# ══════════════════════════════════════════════
# SHEET 5: Raw Data (Power BI source)
# ══════════════════════════════════════════════
ws5 = wb.create_sheet("Raw_Data_PBI")
ws5.freeze_panes = "A2"

pbi_cols = ["Transaction ID", "Date", "Customer ID", "Gender", "Age",
            "Product Category", "Quantity", "Price per Unit", "Total Amount",
            "Agent", "Region", "Month", "Quarter"]
pbi_df = transactions[pbi_cols].copy()
pbi_df["Date"] = pbi_df["Date"].dt.strftime("%Y-%m-%d")

for c, col in enumerate(pbi_cols, 1):
    cell = ws5.cell(1, c, col)
    cell.font = hdr_font(10)
    cell.fill = hdr_fill(C_SUBHDR_BG)
    cell.alignment = center()
    cell.border = border

for r, row in enumerate(pbi_df.itertuples(index=False), 2):
    for c, val in enumerate(row, 1):
        ws5.cell(r, c, val).font = cell_font(9)
        ws5.cell(r, c).border = Border(
            left=Side(style="hair"), right=Side(style="hair"),
            top=Side(style="hair"), bottom=Side(style="hair")
        )

set_col_widths(ws5, [14, 12, 12, 8, 5, 16, 9, 15, 13, 15, 9, 9, 8])

# ══════════════════════════════════════════════
# SHEET 6: KPI Dashboard (summary for PBI)
# ══════════════════════════════════════════════
ws6 = wb.create_sheet("KPI_PBI_Export")

# Monthly aggregated
monthly_agg = monthly_kpi.groupby("Month").agg(
    Total_Sales=("Actual_Sales", "sum"),
    Total_Target=("Target", "sum"),
    Incentive_Old=("Incentive_Old", "sum"),
    Incentive_New=("Incentive_New", "sum"),
    Avg_Pct_Target=("Pct_Target", "mean"),
    Agents_Above_Target=("Pct_Target", lambda x: (x >= 1.0).sum()),
    Agents_Below_80=("Pct_Target", lambda x: (x < 0.8).sum()),
).reset_index()
monthly_agg["ROI_Old"] = (monthly_agg["Total_Sales"] / monthly_agg["Incentive_Old"]).round(2)
monthly_agg["ROI_New"] = (monthly_agg["Total_Sales"] / monthly_agg["Incentive_New"]).round(2)

cols6 = list(monthly_agg.columns)
for c, col in enumerate(cols6, 1):
    cell = ws6.cell(1, c, col)
    cell.font = hdr_font(10)
    cell.fill = hdr_fill(C_SUBHDR_BG)
    cell.alignment = center()
    cell.border = border

for r, row in enumerate(monthly_agg.itertuples(index=False), 2):
    for c, val in enumerate(row, 1):
        cell = ws6.cell(r, c, val)
        cell.font = cell_font(9)
        cell.border = Border(left=Side(style="hair"), right=Side(style="hair"),
                             top=Side(style="hair"), bottom=Side(style="hair"))

set_col_widths(ws6, [10, 14, 14, 14, 14, 14, 18, 16, 10, 10])

wb.save(OUTPUT)
print(f"\n✅ Done! Saved: {OUTPUT}")
print(f"\n📊 Summary:")
print(f"  Transactions processed : {len(df):,}")
print(f"  Agents tracked         : {len(AGENTS)}")
print(f"  Months analyzed        : {len(months)}")
print(f"  Total Sales            : {agent_summary['Total_Sales'].sum():,.0f}")
print(f"  Incentive Cost (Old)   : {agent_summary['Total_Incentive_Old'].sum():,.0f}")
print(f"  Incentive Cost (New)   : {agent_summary['Total_Incentive_New'].sum():,.0f}")
print(f"  ROI Old Scheme         : {monthly_kpi['Actual_Sales'].sum() / monthly_kpi['Incentive_Old'].sum():.2f}x")
print(f"  ROI New Scheme         : {monthly_kpi['Actual_Sales'].sum() / monthly_kpi['Incentive_New'].sum():.2f}x")
